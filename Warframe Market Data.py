#Spencer Oswald

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
path = input("Enter the path to the worksheet\n")
wb = load_workbook(path)
ws = wb.active
partList = []
for i in range(150): #Yes, I realize setting it to max out at 150 is lazy, oh well.
    c = ws["A" + str(i+2)].value
    if c != None:
        partList.append(c)


options = Options()
options.headless = True
driver = webdriver.Firefox(options=options)        
url = "warframe.market/item"
platList = []
for part in partList:
    part = part.replace(' ', '_')
    part = part.lower()
    url = "https://warframe.market/items/"+part
    print (url)
    driver.get(url)
    try:
        plat = driver.find_element_by_xpath("/html/body/section/section/div[2]/section[2]/div[3]/div[2]/div[2]/div/div[1]/div[4]/div/b")
    except:
        print("Oops, maybe a typo in there.")
        plat = "Error"
    else:
        plat = driver.find_element_by_xpath("/html/body/section/section/div[2]/section[2]/div[3]/div[2]/div[2]/div/div[1]/div[4]/div/b")
    if type(plat) != str:
        plat = plat.text
    print (plat)
    platList.append(plat)
driver.close()

i = 0
for plat in platList:
    print(plat)
    ws["C" + str(i+2)] = plat
    i = i + 1

wb.save(path)
wb.close()

